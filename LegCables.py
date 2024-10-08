#Python Search for best cable arrangement: minimum area with prestress and area sizes as design variables
# >>> works best through a driver that runs runLegCables.py see D:\RRD_ENGINEERING\PROJECTS\WE202301_NREL_Brunner\ANALYSIS\Wisdem\blade_driver.py
#      CHECK ALSO BladeOptimizer.py under BladeRunner/wisdem
"""
LegCables.py

Created by Rick Damiani on 2024-08-05.
Copyright (c)  FWTC, LLC. All rights reserved.
"""

#from __future__ import print_function
from cmath import isfinite
import datetime
from tkinter import N
import numpy as np
import argparse
#import copy
from pathlib import Path
import os, time, shutil, copy
import matplotlib.pyplot as plt
#import matplotlib.animation as animation
import openmdao.api as om
#import pandas as pd
import yaml
#import openpyxl

from openmdao.api import IndepVarComp, ExplicitComponent, ExecComp
#from scipy import linalg
from wisdem.commonse import gravity
from wisdem.commonse.cross_sections import Tube
from wisdem.commonse.Material_RRD import Material

class RRDTube(Tube):
    """The old RRD Tube Class was destroyed by NREL, so adding a few back here."""

    def __init__(self, D, t, L=np.NaN, Kbuck=1.0, mat=Material(name='steel',fy=1000.e6)):
        self.mat = mat
        super().__init__( D, t, L, Kbuck)
    @property
    def mass(self):  # mass of tube
        return self.Area*self.L*self.mat.rho
    
class Leg(ExplicitComponent):
   
    def initialize(self):
        self.options.declare('rho_w',types=float, default=1025., desc='sea density')    
        
        self.options.declare('concrete',types=object, desc='material instance for leg',default=Material(matname='concrete'))
        self.options.declare('tendon',types=object, desc='instance of Material object for prestressing tendon; if no prestressing tendons, set=preinf3 to np.zeros(3,1) but provide this material',default=Material(matname='strand'))
        self.options.declare('rebar',types=object, desc='instance of Material object for non-prestressed reinforcement',default=Material(matname='rebar'))
        self.options.declare('tie',types=object, desc='instance of Material object for stirrup/hoop/tie/spiral shear reinforcement',default=Material(matname='rebar'))
        self.options.declare('ld_eff', types=float, desc='distributed load that will be added to distributed gravity load on the leg, given as load/g; e.g.: beff=-rho_w*A_gL to account for buoyancy.  >0 if along gravity;units=N/m/(m/s2)')
        self.options.declare('nlreinf_sets', types=int,  desc='how many sets of longitudinal non-prestressed reinforcement we have got',  default=2)  
        self.options.declare('npreinf_sets', types=int,  desc='how many sets of longitudinal prestressed reinforcement we have got',  default=0)
        self.options.declare('nsreinf_sets', types=int,  desc='how many sets of shear hoop/tie/spirals reinforcement we have got',  default=2)    


    def setup(self):
       
        self.add_input('D_L',  desc='Leg OD', units='m')  
        self.add_input('d_L',  desc='Leg ID', units='m')  
        self.add_input('L_L0', desc='Leg initial (undeflected) length', units='m')
        self.add_input('c',    desc='distance from center hinge on stem to projection of buoyancy can joint onto column', units='m')

        self.add_input('lreinf3', np.zeros(3*self.options['nlreinf_sets']),  desc='triplets for non-prestressed reinforcement [Ns1,As1,Ds1,Ns2,As2,Ds2,...Nsn,Asn,Dsn]')
        """where    Nsi    - int,  number of axial reinforcement elements (nonprestressed) located along circle of Ds1 diameter [-] \n
                    Asi    - float, reinforcement single element area [m2] \n
                    Dsi    - float, circle diameter for reinforcement placement [m] \n"""
        if self.options['npreinf_sets']:
            self.add_input('preinf3',   np.zeros(3*self.options['npreinf_sets']), desc='triplets for prestressed reinforcement [Np1,Ap1,Dp1,...Npn,Apn,Dpn] as for non-prestressed rebar')
        else:
            self.add_input('preinf3',   np.zeros(3), desc='triplets for prestressed reinforcement [Np1,Ap1,Dp1,...Npn,Apn,Dpn] as for non-prestressed rebar')
        self.add_input('sreinf3', np.array([1.,0.,0.]*2), desc='triplets  for shear (ties/hoops/spirals) reinforcement [pv1,Av1,Dv1,pv2,Av2,Dv2,...pvn,Avn,Dvn]')
        """where psi    - float, pitch of shear reinforcement elements  [m] \n
                 Asi    - float, reinforcement element area [m2] \n
                 Dsi    - float, circle diameter for reinforcement placement [m] \n"""
        
        #OUTPUTS
        self.add_output('Tht_L0', desc='Leg undeflected angle with normal to stem''s axis', units='rad')
        #self.add_output('E_c',  desc='Leg concrete Young''s modulus', units='N/m**2')
        self.add_output('A_Lceff',  desc='Leg normalized cross-sectional area', units='m**2')
        self.add_output('Jxx_Lceff',  desc='Leg normalized cross-sectional second area moment of inertia', units='m**2')
        self.add_output('m_L',  desc='Leg distributed mass per unit length', units='kg/m')
        self.add_output('M_L',  desc='Leg total mass including reinforcement steel', units='kg')
        self.add_output('b_eff',  desc='Leg distributed load per g', units='N/m**2/s**2')
        self.add_output('m_eff',  desc='Leg effective mass per g, including m_L, buoyancy, and b_eff', units='N/m**2/s**2')
        self.add_output('A_stot',  desc='Total area of longitudinal non-prestressed reinforcement', units='m**2')
        self.add_output('A_ptot',  desc='Total area of longitudinal prestressed reinforcement', units='m**2')
        self.add_output('A_vtot',  desc='Total area of shear reinforcement per unit leg-span', units='m**2')
        self.add_output('nRC',   desc='Ratio of rebar steel to concrete Young''s moduli', units=None)
        self.add_output('nRCp',  desc='Ratio of tendon steel to concrete Young''s moduli', units=None)
        self.add_output('Vs_shear',  desc='Contribution to shear strength due to shear reinforcement', units='N')

    def compute(self,inputs, outputs):

        tube=Tube( inputs['D_L'], (inputs['D_L']-inputs['d_L'])/2.,inputs['L_L0'], Kbuck=1., mat=self.options['concrete'])

        nRCp=0. #Initialize
        nRC =0. #Initialize
 
        #outputs['E_c']= self.options['concrete'].E #Ec
        outputs['Tht_L0']= np.arcsin(inputs['c']/inputs['L_L0']) # Angle between leg axis and normal to stem's axis, undeflected conditions.
        
        if self.options['rebar'] != None:
            outputs['nRC'] = self.options['rebar'].E/self.options['concrete'].E 
        if self.options['tendon'] != None:
            outputs['nRCp']= self.options['tendon'].E/self.options['concrete'].E 

        #Now find the concrete area properties 
        [Ag,Ac,outputs['A_Lceff'],Aceffp,outputs['A_stot'],outputs['A_ptot'],Jxxg,Jxxc,outputs['Jxx_Lceff'],Jxxstot,Jxxptot,J0g]= \
            AnnAreas(inputs['D_L'], inputs['d_L'], outputs['nRC'],outputs['nRCp'], inputs['preinf3'], inputs['lreinf3'])
        #A_vtot 
        A_vtot=0. #initialize
        spc_avg=0. #initialize
        alpha_avg=0. #initialize

        for iv in range(self.options['nsreinf_sets']):
            spc= inputs['sreinf3'][iv*3+0] 
            D_plc= inputs['sreinf3'][iv*3+2] 
            A_tie= inputs['sreinf3'][iv*3+1] 
            alpha_tie= np.arctan(np.pi*D_plc/spc)
            A_vtot += A_tie 
            spc_avg += spc/self.options['nsreinf_sets'] 
            alpha_avg += alpha_tie/self.options['nsreinf_sets'] 

        outputs['A_vtot']=A_vtot #single tie leg
        #outputs['alpha_avg']= alpha_avg

        #and the mass per unit length
        options=self.options
        outputs['m_L']= XsectMass(options['concrete'],options['rebar'],options['tie'],options['tendon'], inputs['D_L'], inputs['d_L'], inputs['lreinf3'], inputs['sreinf3'], inputs['preinf3'])
        outputs['M_L']= outputs['m_L'] * inputs['L_L0']
        outputs['b_eff']= self.options['ld_eff'] - self.options['rho_w']*Ag
        outputs['m_eff']= outputs['m_L'] + outputs['b_eff']
        
        #shear strength due to reinforcement Vs_shear(bw, d, s, Ast, fyst, fyc, alpha=90.
        outputs['Vs_shear'] = Vs_shear(inputs['D_L'],0.8*inputs['D_L'] , spc_avg, A_vtot, options['tie'].fy, options['concrete'].fyc, alpha=np.rad2deg(alpha_avg))

class Cable:
    def __init__(self, D, xyz, sig0, uvw, material=Material(name='steel',fy=1000.e6)):
        """TODO add gravity effects

        Parameters
        ----------
        D : float [m] OD. \n
        theta : float [deg], angle of leg to the horizontal, positive. \n
        xyz: float(2,3) [m], undisturbed nominal geoemtric length, first row coordinates of end A, second row coordinates of end B. \n
        sig0: float [Pa], prestress. \n
        uvw: float(3) [N], displacement of end B, after postensioning. \n
        N: float [N], actual internal load (including pretension). \n
        material: object of class Material_RRD
        """
        #Unpack        
        self.D = D
        
        self.material = material
        self.rho = material.rho
        self.Lg0 = np.linalg.norm(xyz[1,:]-xyz[0,:]) #geometric length
        #self.Lg0 = np.sqrt( ((xyz[1,:]-xyz[0,:])**2).sum())
        #print('past Lg0')
        self.sig0=sig0
        
        self.tube=RRDTube( self.D, self.D/2., self.Lg0)
        self.uvw = uvw
        self.xyz = xyz
        #calculate some outputs
    @property
    def A_xsect(self): 
        """Cross-sectional Area"""
        return self.tube.Area  #cross sectional area
    
    @property
    def L0(self):
        """Unstretched length"""
        return self.Lg0/(1.+self.eps0)

    @property
    def eps0(self):
        """Strain only due to pretension"""
        return self.sig0/self.material.E

    @property
    def Bp(self):
        """Final coordinates of end B flaot(3) [m]"""
        return self.xyz[1,:] + self.uvw

    @property
    def strain_tension(self):
        """Additional strain due to displacement of end B, it will add to pretension; and final load components at end B;\n
        OUTPUT:\n
        Delta_eps: float, extra strain due to final displacement [-]\n
        Nxyz:      float(3), components of the axial load along X,Y,Z at end A [N]\n
        N:         float, axial load magnitude [N]\n
        Lf:        float, final length of the cable [m]. """

        ABp = self.Bp - self.xyz[0,:] #final vector 
        Lf = np.linalg.norm(ABp) #final cable length np.sqrt((ABp**2).sum()) 
        #print("post Lf calc")
        Delta_eps = (Lf-self.Lg0)/self.L0 #this is the additional eps from eps0, because of the final u,v,w displacements of end B
        
        Nxyz = ABp/Lf * (self.N0 + self.material.E * self.A_xsect * Delta_eps) #components of final loads along, X,Y,Z
        if self.eps0+Delta_eps <0.: #no compression for cables 
            Nxyz=np.zeros(3)    

        N = np.linalg.norm(Nxyz)
        return Delta_eps, Nxyz, N, Lf

    @property
    def N0(self):
        """Pretension Load"""
        return self.material.E * self.A_xsect * self.eps0

    @property
    def mass(self):
        """Dry Mass"""
        return self.tube.mass  #dry mass

    @property
    def sig(self):
        """Effective stress given axial load"""
        N = self.strain_tension[2]
        return N/self.A_xsect

    def plot_cable(self):        
        """Plot cable"""
        #fig= plt.figure()
        ax = plt.axes(projection='3d')

        ax.plot(self.xyz[:,0],self.xyz[:,1],self.xyz[:,2],label='D={:5.3f} m L= {:5.2f} m'.format(*self.D,self.Lg0),color='red')

        #lower_points = self.xyz[0,:]
                        
                        #0.5*self.D* [-sintht,costht], 
                      #- 0.5*self.D* [-sintht,costht]]

        #upper_point0 = self.Lg0 * [costht,sintht]  
        #upper_points = upper_point0*[ 1. + lower_points]

        #plt.fill(lower_points,upper_points, color=plt.cm.Dark2(loc_matidx[it]), label=cable)
        
        plt.axis('equal')
        plt.grid()
        #plt.title('Cable Geometry:  D={:5.3f} m L= {:5.2f} m '.format(*self.D,self.Lg0))
        #plt.plot(upper_points[0,:],upper_points[1,:],color='black',linestyle=':')
        #plt.legend(myhandles, mylabels)    
        ax.set_xlabel('x_R[m]')
        ax.set_ylabel('y_R[m]')
        ax.set_zlabel('z_R[m]')
        ax.legend()
        #plt.annotate('$\mu_s$({:3.2},{:3.2})= {:5.3e}'.format(*crit_points[ii,:], eps_z[var][ii]*10**6), xy=np.flip(crit_points[ii,:]), xycoords='data',\
        #                       xytext=(0, 10*(ii+1)*(-1)**i_surf), textcoords='offset points', arrowprops=dict(arrowstyle="->", connectionstyle="arc3"),  horizontalalignment='center', verticalalignment='bottom')
            
        #plt.show()            
 
class legCable(om.ExplicitComponent) :   
    """Compute the main properties of a cable.\n
       """
    def initialize(self):
        self.options.declare("material")
        self.options.declare("xyz0",  default=np.zeros([2,3]), desc="Cable [end A; end B] coordinates at setup [m]" ) 
        self.options.declare("z_legroot",  default=0., desc="Leg root elevation [m]" ) 
        #Next 2 are for guesses
        self.options.declare("sig00",  default=50.e6, desc="Cable prestress initial guess [Pa]" ) 
        self.options.declare("D_C00",      default=0.1, desc="Cable D initial guess [Pa]" ) 
        self.count4plot =0 

    def setup(self):
        self.xyz0     = self.options["xyz0"] 
        self.material = self.options["material"]
        
        self.add_input("sig0",      val=0., desc="Pretension stress"                    , units="N/m**2"     )
        self.add_input("uvwB",      val=np.zeros(3), desc="Cable end B, final displacement" , units="m"     )
        self.add_input("D",         val=0., desc="Cable diameter", units="m")
        
        #Calculate initial guesses for Nxyz
        mycable = Cable(self.options['D_C00'], self.xyz0, self.options["sig00"], np.zeros(3), material=self.material)
        Nxyz_guess =-mycable.strain_tension[1]
        
        self.add_output("Nxyz",     val=Nxyz_guess,   desc="Final load components along x,y,z at end B, after pretension and extra deformation", units="N")
        self.add_output("Bp",       val=self.xyz0[1,:], desc="Final x,y,z of end B, after pretension and extra deformation,  absolute", units="m")
        self.add_output("Bprel",    val=self.xyz0[1,:]-np.array([0.,0.,self.options['z_legroot']]), desc="Final x,y,z of end B, after pretension and extra deformation,  relative to leg root", units="m")
        self.add_output("sig",      val=0., desc="Actual maximum axial stress"  , units="Pa"     )
        self.add_output("mass",     val=0., desc="Dry mass"  , units="kg"     )

    def setup_partials(self):
        self.declare_partials('*', '*', method='fd')

    def compute(self,inputs, outputs,discrete_inputs=None, discrete_outputs=None):        
        """Find the axial load, the strain"""
        mycable = Cable(inputs['D'], self.xyz0, inputs["sig0"], inputs["uvwB"], material=self.material)
        #if self.count4plot==0:
        #    fig= plt.figure()

       # self.count4plot +=1

        outputs["Nxyz"]  = -mycable.strain_tension[1] #note "-" because it is at end B
        outputs['mass']  = mycable.mass
        outputs['sig']   = mycable.sig
        outputs['Bp']    = mycable.Bp 
        outputs['Bprel'] = outputs['Bp'] - np.array([0.,0.,self.options["z_legroot"]])
        #mycable.plot_cable()
       # if self.count4plot==3:
        #    plt.show()
        

class uvwProcessor(om.ExplicitComponent) :   
    """Computes uvwB from uvw_legtip and theta_legtip.\n
       """
    def initialize(self):
        self.options.declare("C1orC2", types=int,   default=1,  desc="C1 (=-1) or C2 (=+1), i.e. negative(1) y or positive(1) y cable") 
        self.options.declare("D_L",  types=float, default=0., desc="Leg OD" ) 
        self.options.declare('L_L0', types=float, default=0., desc='Leg initial (undeflected) length')
        self.options.declare('Lh',   types=float, default=0., desc='distance from root to cable hinge along axis')

    def setup(self):
        self.D_L  = self.options["D_L"]
        self.L_L0 = self.options["L_L0"]
        self.Lh  = self.options["Lh"]
        self.C1orC2 = self.options["C1orC2"] 

        self.add_input("uvw_legtip",     val=np.zeros(3), desc="Leg tip final displacement" , units="m"     )
        self.add_input("tht_legtip",   val=0., desc="Leg tip final rotation about its axis (twist)" , units="rad"     )
        
        self.add_output("uvwB",          val=np.zeros(3), desc="Cable end B final displacement" , units="m"     )

    def setup_partials(self):
        self.declare_partials('uvwB', 'uvw_legtip', method='fd')
        self.declare_partials('uvwB', 'tht_legtip', method='fd')

    def compute(self,inputs, outputs, discrete_inputs=None,discrete_outputs=None):        
        """Find the displacement of end B of the cable"""
       
        L_L0       = self.L_L0
        D_L        = self.D_L
        Lh         = self.Lh
        C1orC2     = self.C1orC2
        uvw_legtip = inputs["uvw_legtip"]
        theta      = inputs["tht_legtip"]
        
        delta_tvec = uvw_legtip/L_L0
    
        junk=uvw_legtip[1]/L_L0
        cost=np.cos(theta)
        
        delta_nvec = np.array( [-cost * junk  ,cost*(1-junk**2) -1, np.sin(theta)] ).T

        
        #next is CC' and DD'
        CCp =  Lh*delta_tvec + C1orC2 * D_L/2.*delta_nvec
        
        outputs['uvwB']  = CCp
        pass
    
class LegCableBal(om.ImplicitComponent) :   
    """Compute the axial loads in the cables, leg, as well as tip deflections.\n
    """
    def initialize(self):
        self.options.declare("parameters")
        self.options.declare("Fxyz_tip",types=float, default=np.zeros(6), desc="Leg tip loads, forces and moments in global CS [N and Nm]" )
    def setup(self):
        parameters = self.parameters = self.options['parameters']
        
        self.Fxyz_tip = self.options['Fxyz_tip'][0:3] #"3 forces at leg tip"      , units="N"     )
        self.Mxyz_tip = self.options['Fxyz_tip'][3:6] #"3 moments at leg tip"      , units="N*m"     )
        
        self.abc  = parameters['abc'] #,     val=np.zeros(3), desc='Geometry a,b,c paramaters, a=keel to leg-root hinge, c=leg-tip to leg-root vertical distance, b= top cable hinge to leg root vertical distance', units='m')
        self.D_L  = parameters['D_L']  #Leg OD
        self.L_L0 = parameters['L_L0']  #Leg length
        self.L_Lxyz0 = parameters['L_Lxyz0']  #Leg AB components in the glbal CS
        self.M_L  = parameters['M_L']   #Leg Mass
        L_lh = parameters["Lhs"][0] #distance from root of leg to lower cable hinges
        L_uh = parameters["Lhs"][1] #distance from root of leg to upper cable hinges

        self.add_input( "Nxyz_UC1",   val=np.zeros(3), desc="Upper C1 cable components of tension at end B",       units="N"   )
        self.add_input( "Nxyz_UC2",   val=np.zeros(3), desc="Upper C2 cable components of tension at end B",       units="N"   )
        self.add_input( "Nxyz_LC1",   val=np.zeros(3), desc="Lower C1 cable components of tension at end B",       units="N"   )
        self.add_input( "Nxyz_LC2",   val=np.zeros(3), desc="Lower C2 cable components of tension at end B",       units="N"   )

        self.add_input( "UC1_Bprel",     val=np.array([L_uh,-self.D_L/2.,0.]), desc="Upper C1 cable final coordinates of end B relative to leg root",       units="m"   )
        self.add_input( "UC2_Bprel",     val=np.array([L_uh,self.D_L/2., 0.]), desc="Upper C2 cable final coordinates of end B relative to leg root",       units="m"   )
        self.add_input( "LC1_Bprel",     val=np.array([L_lh,-self.D_L/2.,0.]), desc="Lower C1 cable final coordinates of end B relative to leg root",       units="m"   )
        self.add_input( "LC2_Bprel",     val=np.array([L_lh,self.D_L/2., 0.]), desc="Lower C2 cable final coordinates of end B relative to leg root",       units="m"   )
        self.add_input( "uvw_legtip_guess", val=np.zeros(3), desc="guess u,v, w at leg tip", units="m"   )
        self.add_input( "tht_legtip_guess", val=0., desc="guess theta  at leg tip", units="rad"   )
        self.add_input( "XYZ0_guess",       val=np.zeros(3), desc="guess XYZ0", units="N"   )

        self.add_output( "uvw_legtip", val=np.zeros(3), desc="u,v, w at leg tip", units="m"   )
        self.add_output( "tht_legtip", val=0., desc="twist at leg tip", units="rad"   )
        self.add_output( "XYZ0",       val=np.zeros(3), desc="Leg root reactions",       units="N"   )

    def setup_partials(self):
        self.declare_partials(of='*', wrt='*',method='fd')
   
    def apply_nonlinear(self, inputs, outputs, residuals):
    #Initialize some local vars shortening names
        #parameters
        L_L0     =self.L_L0
        L_Lxyz0  =self.L_Lxyz0

        M_L      =self.M_L

        Fxyz =self.Fxyz_tip
        Mxyz =self.Mxyz_tip
        #outputs
        XYZ0         = outputs["XYZ0"]
        uvw_legtip   = outputs["uvw_legtip"]
        theta        = outputs["tht_legtip"]

        #Cable data
        Nxyz_UC1 = inputs['Nxyz_UC1']
        Nxyz_UC2 = inputs['Nxyz_UC2']
        Nxyz_LC1 = inputs['Nxyz_LC1']
        Nxyz_LC2 = inputs['Nxyz_LC2']

        UC1_Bp = inputs["UC1_Bprel"]
        UC2_Bp = inputs["UC2_Bprel"]
        LC1_Bp = inputs["LC1_Bprel"]
        LC2_Bp = inputs["LC2_Bprel"]
        
        L_L = L_Lxyz0+uvw_legtip #final leg vector AB
        #>>>>>>>System of equations<<<<< unknowns (outputs) uvw_legtip theta_legtip
        #rigid leg condition:
        residuals['uvw_legtip'][0]  = L_L0**2 - (L_L**2).sum() #((L_L0  + uvw_legtip[0])**2 + uvw_legtip[1]**2 + uvw_legtip[2]**2)  #length of leg is invariant   ::Eq#1
       # constitutive equation for cables are now in the cable component
# 
 #       6 DOF static balance, moments about leg root (no torsional hinge)
        residuals['XYZ0'][0] = Fxyz[0] + XYZ0[0] + Nxyz_UC1[0] + Nxyz_UC2[0] + Nxyz_LC1[0] + Nxyz_LC2[0] #Balance along X                ::Eq#2  
        residuals['XYZ0'][1] = Fxyz[1] + XYZ0[1] + Nxyz_UC1[1] + Nxyz_UC2[1] + Nxyz_LC1[1] + Nxyz_LC2[1] #Balance along Y                ::Eq#3  
        residuals['XYZ0'][2] = Fxyz[2] + XYZ0[2] + Nxyz_UC1[2] + Nxyz_UC2[2] + Nxyz_LC1[2] + Nxyz_LC2[2] - M_L*gravity #Balance along Z  ::Eq#4  
        # 
        residuals['tht_legtip'] = Mxyz[0] - L_L[2]*Fxyz[1] + (Fxyz[2] - M_L * gravity/2.) * L_L[1]          + \
                                                      UC1_Bp[1]*Nxyz_UC1[2] - UC1_Bp[2]*Nxyz_UC1[1]  \
                                                    + UC2_Bp[1]*Nxyz_UC2[2] - UC2_Bp[2]*Nxyz_UC2[1]  \
                                                    + LC1_Bp[1]*Nxyz_LC1[2] - LC1_Bp[2]*Nxyz_LC1[1]  \
                                                    + LC2_Bp[1]*Nxyz_LC2[2] - LC2_Bp[2]*Nxyz_LC2[1]   #Eq.#5
# 
        residuals['uvw_legtip'][2] = Mxyz[1] + L_L[2]*Fxyz[0] + (-Fxyz[2] + M_L * gravity/2.) * L_L[0]          + \
                                                      - UC1_Bp[0]*Nxyz_UC1[2] + UC1_Bp[2]*Nxyz_UC1[0]  \
                                                      - UC2_Bp[0]*Nxyz_UC2[2] + UC2_Bp[2]*Nxyz_UC2[0]  \
                                                      - LC1_Bp[0]*Nxyz_LC1[2] + LC1_Bp[2]*Nxyz_LC1[0]  \
                                                      - LC2_Bp[0]*Nxyz_LC2[2] + LC2_Bp[2]*Nxyz_LC2[0]  #Eq.#6
# 
        residuals['uvw_legtip'][1] = Mxyz[2] + L_L[0]*Fxyz[1] - L_L[1]*Fxyz[0]                                  + \
                                                      - UC1_Bp[1]*Nxyz_UC1[0] + UC1_Bp[0]*Nxyz_UC1[1]  \
                                                      - UC2_Bp[1]*Nxyz_UC2[0] + UC2_Bp[0]*Nxyz_UC2[1]  \
                                                      - LC1_Bp[1]*Nxyz_LC1[0] + LC1_Bp[0]*Nxyz_LC1[1]  \
                                                      - LC2_Bp[1]*Nxyz_LC2[0] + LC2_Bp[0]*Nxyz_LC2[1]  #Eq.#7

        #print( " LegCableBal eq",residuals)
        outs= residuals
        #return outs

    def guess_nonlinear(self, inputs, outputs, residuals):
        #Now set the initial guesses
        if any(inputs['uvw_legtip_guess']):
            outputs['uvw_legtip']= inputs['uvw_legtip_guess']
        else:   
            outputs['uvw_legtip']= [-2.3011e-003, 0.,0.41364] #[-np.cos(2/180.*np.pi),0.0,np.sin(2/180.*np.pi)] 
        if any(inputs['tht_legtip_guess']):
            outputs['tht_legtip']= inputs['tht_legtip_guess']
        else:   
            outputs['tht_legtip']= 0.0
        if any(inputs['XYZ0_guess']):
            outputs['XYZ0']= inputs['XYZ0_guess']
        else:   
            outputs['XYZ0']=  np.array([1.7273e+007,0,-2.2655e+006]) #self.M_L*gravity/2.])



    def solve_nonlinear(self, inputs, outputs, residuals):
        pass
        #solution = self.nonlinear_solver(residuals,iprint=3)

    def compute(self,inputs, outputs, discrete_inputs=None,discrete_outputs=None):
        #output['tht_legtip']=
        pass
       # Fxyz_tip = inputs["Fxyz_tip"]
       # Mxyz_tip = inputs["Mxyz_tip"]
       # L_L0 = self.LL0


#AUXILIARY FUNCTIONS       

class cycle(om.Group):
    """ Group containing the components that form the coupled "cycle" to be solved first before optimizer can do its thing"""
    def initialize(self):
        self.options.declare("parameters")
        self.options.declare("Fxyz_tip",types=float, default=np.zeros(6), desc="Leg tip loads, forces and moments in global CS [N and Nm]" )
        



class CableGroup(om.Group):
    """ Group containing the components to passs the loads, get blade xsec properties and find strains"""
    def initialize(self):
        self.options.declare("parameters")
        self.options.declare('parallel_derivs', False, types=bool, allow_none=True)

    def setup(self):
        parameters = self.options["parameters"]
        n_DLCs = parameters['n_DLCs'] #number of load cases
        parallel_derivs = self.options['parallel_derivs']
        
        #Add subsystems
        self.add_subsystem("D_Cs", IndepVarComp("D_C",val=np.zeros(2),units='m')) #lower and upper cable diameters
        self.add_subsystem("sig0_Cs", IndepVarComp("sig0_C",val=np.zeros(2),units="Pa")) #lower and upper cable prestresses
        
        #_____Cycle: Create a subgroup to solve the coupled system
        cycle = self.add_subsystem("cycle",om.Group(),promotes=['*'])

        cycle.add_subsystem("uvwLC1", uvwProcessor(C1orC2=-1,D_L=parameters["D_L"],L_L0=parameters["L_L0"],Lh=parameters["Lhs"][0]),promotes_inputs=["uvw_legtip","tht_legtip"])
        cycle.add_subsystem("uvwLC2", uvwProcessor(C1orC2=+1,D_L=parameters["D_L"],L_L0=parameters["L_L0"],Lh=parameters["Lhs"][0]),promotes_inputs=["uvw_legtip","tht_legtip"])
        
        cycle.add_subsystem("uvwUC1", uvwProcessor(C1orC2=-1,D_L=parameters["D_L"],L_L0=parameters["L_L0"],Lh=parameters["Lhs"][1]),promotes_inputs=["uvw_legtip","tht_legtip"] )
        cycle.add_subsystem("uvwUC2", uvwProcessor(C1orC2=+1,D_L=parameters["D_L"],L_L0=parameters["L_L0"],Lh=parameters["Lhs"][1]),promotes_inputs=["uvw_legtip","tht_legtip"])

        cycle.add_subsystem("UC1", legCable(xyz0=parameters['UC1_xyz0'],material=parameters["UCmat"],sig00=parameters["sig00"][1],D_C00=parameters["D_C00"][1],z_legroot=parameters['z_legroot']))
        cycle.add_subsystem("LC1", legCable(xyz0=parameters['LC1_xyz0'],material=parameters["LCmat"],sig00=parameters["sig00"][0],D_C00=parameters["D_C00"][0],z_legroot=parameters['z_legroot']))
        #Symmetric cables provide symmetric geometry in mirrored y
        junk = np.copy(parameters["UC1_xyz0"])
        junk[:,1] *=-1
        cycle.add_subsystem("UC2", legCable(xyz0=junk,material=parameters["UCmat"],sig00=parameters["sig00"][1],D_C00=parameters["D_C00"][1],z_legroot=parameters['z_legroot']))
        junk = np.copy(parameters["LC1_xyz0"])
        junk[:,1] *=-1
        cycle.add_subsystem("LC2", legCable(xyz0=junk,material=parameters["LCmat"],sig00=parameters["sig00"][0],D_C00=parameters["D_C00"][0],z_legroot=parameters['z_legroot']))
        
        cycle.add_subsystem('LegCableBal',LegCableBal(parameters=parameters,Fxyz_tip=),promotes_outputs=["uvw_legtip","tht_legtip"]) 
        cycle.connect("uvwUC1.uvwB","UC1.uvwB")
        cycle.connect("uvwUC2.uvwB","UC2.uvwB")
        cycle.connect("uvwLC1.uvwB","LC1.uvwB")
        cycle.connect("uvwLC2.uvwB","LC2.uvwB")
        #connections to balance component
        
        cycle.connect("UC1.Nxyz", "LegCableBal.Nxyz_UC1")
        cycle.connect("UC2.Nxyz", "LegCableBal.Nxyz_UC2")
        cycle.connect("LC1.Nxyz", "LegCableBal.Nxyz_LC1")
        cycle.connect("LC2.Nxyz", "LegCableBal.Nxyz_LC2")
        cycle.connect("UC1.Bprel","LegCableBal.UC1_Bprel")
        cycle.connect("UC2.Bprel","LegCableBal.UC2_Bprel")
        cycle.connect("LC1.Bprel","LegCableBal.LC1_Bprel")
        cycle.connect("LC2.Bprel","LegCableBal.LC2_Bprel")

        cycle.nonlinear_solver = om.NonlinearBlockGS()
        #_____________end cycle group
        
        self.add_subsystem("UC1LC1_mass", ExecComp("mass=m_UC + m_LC", mass={'units':"kg"}, m_UC={'val':1000.,'units':"kg"},m_LC={'val':1000.,'units':"kg"}) )#lower and upper cable ODs
        self.add_subsystem("cable_totD", ExecComp("totD=D_UC + D_LC", totD={'units':"m"},D_UC={'val':0.2,'units':"m"},D_LC={'val':0.2,'units':"m"}) )#lower and upper cable ODs        
        
        #Note: because promote=["*"] then no need for "wt_init." before "blade" or "airfoils" etc.)

        #connections to cable components
        self.connect("D_Cs.D_C",["UC1.D","UC2.D","cable_totD.D_UC"], src_indices=1)
        self.connect("D_Cs.D_C",["LC1.D","LC2.D","cable_totD.D_LC"], src_indices=0)
        self.connect("UC1.mass","UC1LC1_mass.m_UC")
        self.connect("LC1.mass","UC1LC1_mass.m_LC")

        self.connect("sig0_Cs.sig0_C",["UC1.sig0","UC2.sig0"], src_indices=1)
        self.connect("sig0_Cs.sig0_C",["LC1.sig0","LC2.sig0"], src_indices=0)


        cycle.nonlinear_solver = om.NewtonSolver(solve_subsystems=False)
        cycle.nonlinear_solver.options['iprint'] = 3
        cycle.nonlinear_solver.options['maxiter'] = 20# 20
        cycle.nonlinear_solver.options['debug_print'] = True# debug
        cycle.nonlinear_solver.options['err_on_non_converge'] = False# debug
        cycle.linear_solver = om.DirectSolver()
        

        # Parallel Subsystem for load cases.
        par = self.add_subsystem('parallel', om.ParallelGroup())      
        # Determine how to split cases up over the available procs.
        nprocs = self.comm.size
        divide = divide_cases(n_DLCs, nprocs) 
        for j, this_proc in enumerate(divide):
            num_rhs = len(this_proc)

            name = 'sub_%d' % j
            sub = par.add_subsystem(name, cycle)
            Fxyz_tip = parameters["Fxyz_tip"][j,:]

        #plot_legcable(np.array([parameters['LC1_xyz0'],parameters['LC1_xyz0']*[1,-1,1],parameters['UC1_xyz0'],parameters['UC1_xyz0']*[1,-1,1]]),)

    def compute(self,inputs, outputs):
      """compute cable OD and sig0 to satisfy constraints"""
      
      pass
        
def main(**kwargs):
    #Note coordinates here are brought back to (0,0,0), we should really stick to absolute CS
    #initialize:

    D_c      = kwargs.get('D_c',None)
    sig0_c   = kwargs.get('sig0_c',None)
    
   
    D_bounds    = kwargs.get('D_bounds',None)    #(n_cables,2) min/max value ODs of each cables to optimize 
    sig0_bounds = kwargs.get('sig0_bounds',None) #(n_cables,2) min/max value sig0 of each cables to optimize 
   
    parameters= {} #iniitalize 
    #Read in the baseline input yaml for RAFT component, and pass that to the problem as options, which avoid reading it in if multiple components are called that use it
    #bline_yaml    = kwargs['bline_yaml']
    #bline = readinput(bline_yaml)
    infile=kwargs.get('xls_input',None) 
    if not( infile== None):
        parameters['LC1_xyz0'],parameters['UC1_xyz0'],[parameters['D_L'],parameters['d_L'],parameters['L_L0']],parameters['M_L'],\
           parameters['abc'],[parameters['D_S'],parameters['d_S'],parameters['L_S'],parameters['draft_S']]     = read_inputs(infile)
        #the following 2 assume Leg horizontal, might change in the future
        parameters['L_Lxyz0']=np.array([parameters['L_L0'],0.,parameters['abc'][2]],dtype=float).squeeze()
        parameters['z_legroot']= parameters['draft_S']+parameters['abc'][0]
        #now reset coordinates so (0,0,0) is the leg root
        reset_x= parameters['D_S']/2.
        parameters['LC1_xyz0'][:,0] -= reset_x
        parameters['UC1_xyz0'][:,0] -= reset_x

    else:
        parameters['LC1_xyz0'] = kwargs['LC1_xyz0'] 
        parameters['UC1_xyz0'] = kwargs['UC1_xyz0']                 
        parameters['D_L']      = kwargs['D_L']
        parameters['L_L0']     = kwargs['L_L0']
        parameters['M_L']      = kwargs['M_L']
        parameters['abc']      = kwargs['abc']
        parameters['L_Lxyz0']  = kwargs['L_Lxyz0']
        parameters['z_legroot']= kwargs['z_legroot']
        
        #parameters['zoff'] = parameters['L_L0']-parameters['Lhs']
    
    parameters['Lhs']  = np.array([parameters['LC1_xyz0'][1,0],parameters['UC1_xyz0'][1,0]])
    
    #Optimizer yes or no
    opti                   = kwargs['opti']
    
    
    parameters['D_C00']      = D_c #for guess
    parameters['sig00']    = sig0_c #for guess
    
    parameters['UCmat']    = kwargs['UCmat']
    parameters['LCmat']    = kwargs['LCmat']
    parameters['PSF_mat']  = kwargs['PSF_mat']
    parameters['Fxyz_tip'] = kwargs['Fxyz_tip']
    parameters['n_DLCs']   = parameters['Fxyz_tip'].shape[0] 

    #___________________________________#
    
    prob= om.Problem( model=CableGroup(parameters=parameters)) 
          
    if opti:
        #prob.driver = om.ScipyOptimizeDriver()
        #prob.driver.options['optimizer'] =  #'COBYLA' #'SLSQP'
        SNOPT=False

        prob.driver = om.ScipyOptimizeDriver()
        prob.driver.options['optimizer'] =  'COBYLA' #'SLSQP'#'COBYLA' #'COBYLA' #'SLSQP' #'COBYLA' #'SLSQP' 

        prob.driver.options['maxiter'] = 300 #ScipyOpt
        prob.driver.options['tol'] = 1e-3 #1e-8 ScipyOpt

        if SNOPT:
            prob.driver =om.pyOptSparseDriver()
            prob.driver.options['optimizer'] = "SNOPT" #"SLSQP" #"SNOPT" #'SNOPT'
            # prob.driver.options['maxiter'] = 100
            #prob.driver.options['tol'] = 1e-3 #1e-8
            prob.driver.opt_settings={'Major feasibility tolerance': 1e-3,\
                'Minor feasibility tolerance': 1e-3,\
                'Major optimality tolerance': 1e-3,\
                'Function precision': 1e-3}
        else:
            prob.driver.opt_settings={'MAXIT': 400,'ACC': 1e-3  }
            
        #DESIGN VARS

        prob.model.add_design_var("D_Cs.D_C",     lower=D_bounds[0,0], upper=D_bounds[0,1], ref= 0.5)   #indices=0,
        #prob.model.add_design_var("LC1.D",     lower=D_bounds[1,0], upper=D_bounds[1,1])   #indices=0,
        prob.model.add_design_var("sig0_Cs.sig0_C",  lower=sig0_bounds[0], upper=sig0_bounds[1], ref=1000.e6) 

        #OBJECTIVE 
        #prob.model.add_objective("cable_totD.totD", ref =0.5)#,  ref=5000.) #this means reducing structural mass
        prob.model.add_objective("UC1LC1_mass.mass", ref =10000.)#,  ref=5000.) #this means reducing structural mass
        
        #CONSTRAINTS  
        prob.model.add_constraint('UC1.sig',  lower=10.e6, upper=parameters['UCmat'].fp/parameters['PSF_mat'], ref=1000.e6) #This is yield constraint and no slacke.g., 10 MPa, to avoid slack conditions
        prob.model.add_constraint('UC2.sig',  lower=10.e6, upper=parameters['UCmat'].fp/parameters['PSF_mat'], ref=1000.e6) #This is yield constraint and no slack
        prob.model.add_constraint('LC1.sig',  upper=parameters['LCmat'].fp/parameters['PSF_mat'], ref=1000.e6) #This is yield constraint
        prob.model.add_constraint('LC2.sig',  upper=parameters['LCmat'].fp/parameters['PSF_mat'], ref=1000.e6) #This is yield constraint
        
        

        prob.model.add_constraint('uvw_legtip', indices=range(1,3),  upper=np.array([1.5,1.5]), ref=1.5) #This is max deflection of leg tip in y and z
        
        #RECORDER
        recorder = om.SqliteRecorder('cases.sql') #creates a recorder variable
        # Attach a recorder to the problem
        prob.add_recorder(recorder)

    prob.setup()
    
    # Load initial wind turbine data from wt_initial to the openmdao problem; Order of these two has been swapped, following runBladeOptimizer
    #wt_opt= yaml2openmdao(wt_opt, modeling_options, wt_init, analysis_options) 
            
    #Here are the only real inputs for now:
       
    prob.set_val('D_Cs.D_C',D_c)
    prob.set_val('sig0_Cs.sig0_C',sig0_c)
    
    #_____________________________#

    prob.run_model()
    
    #Print some outputs
    print_results(prob,post_opt=False, bline_yaml=None)

    ax0=plot_legcable(np.array([parameters['LC1_xyz0'],parameters['LC1_xyz0']*[1,-1,1],parameters['UC1_xyz0'],parameters['UC1_xyz0']*[1,-1,1]]),\
                  D_cables=np.tile(prob.get_val('D_Cs.D_C'),[2,1]).T.reshape([1,-1]).squeeze())
    
    prob.set_val('LegCableBal.uvw_legtip_guess',prob.get_val('uvw_legtip')) #Assign new guesses
    prob.set_val('LegCableBal.tht_legtip_guess',prob.get_val('tht_legtip')) #Assign new guesses
    prob.set_val('LegCableBal.XYZ0_guess',prob.get_val('LegCableBal.XYZ0')) #Assign new guesses

    if opti:
        prob.set_solver_print(level=1)
        prob.model.approx_totals()
    
        prob.run_driver()
    
        prob.record("after_run_driver")
        # Instantiate your CaseReader
        cr = om.CaseReader("cases.sql")
        # Isolate "problem" as your source
        driver_cases = cr.list_cases('problem')
        # Get the first case from the recorder
        case = cr.get_case('after_run_driver')
        # get_val can convert your result's units if desired
        #const_K = case.get_val("con1", units='K')
        # list_outputs will list your model's outputs and return a list of them too
        #print(case.list_outputs())
        objectives = case.get_objectives()
        design_vars = case.get_design_vars()
        constraints = case.get_constraints()
    
        #PRINT RESULTS
        
        #print('constraints uvw_legtip[1,2] [m] \t',constraints['uvw_legtip'])
        prob.list_problem_vars(cons_opts=['lower','upper'],desvar_opts=['lower','upper'],print_arrays=True)
        

        #prob.model.list_inputs(print_arrays=True,units=True)
        #prob.model.list_outputs(print_arrays=True,units=True)
        #prob.list_problem_vars(print_arrays=True)

        print('______________________________________________________\n')
        print('Optimized cable mass ={:5.3f} kg'.format(*prob.get_val('UC1LC1_mass.mass')))
        #print('D_c=',design_vars['D_Cs.D_C'])
        #print('sig0_c=',design_vars['sig0_Cs.sig0_C'])
        print_results(prob,post_opt=True, bline_yaml=None)
        #Final cable configuration
        LC1_xyz = np.copy(parameters['LC1_xyz0'])
        LC1_xyz[1,:] = prob.get_val('LC1.Bp')
        LC2_xyz = np.copy(LC1_xyz)*np.array([[1.,-1.,1.],[1.,-1.,1.]])
        LC2_xyz[1,:] = prob.get_val('LC2.Bp')
        UC1_xyz = np.copy(parameters['UC1_xyz0'])
        UC1_xyz[1,:] = prob.get_val('UC1.Bp')
        UC2_xyz = np.copy(UC1_xyz)*np.array([[1.,-1.,1.],[1.,-1.,1.]])
        UC2_xyz[1,:] = prob.get_val('UC2.Bp')
        
        plot_legcable(np.array([parameters['LC1_xyz0'],parameters['LC1_xyz0']*[1,-1,1],parameters['UC1_xyz0'],parameters['UC1_xyz0']*[1,-1,1]]),\
                      np.array([LC1_xyz,LC2_xyz,UC1_xyz,UC2_xyz]), D_cables=np.tile(prob.get_val('D_Cs.D_C'),[2,1]).T.reshape([1,-1]).squeeze(),ax=ax0)
        pass
#___________________________________________#
#Auxiliary functions
def _nvec (tx,ty,tz,theta):
    """This function finds [nx,ny,nz] unit vector normal to another vector [tx,ty,tz], and assumes an angle theta around the t vector.\n
    INPUTS:\n
    tx,ty,tz: float, components of the first vector to find a normal to. [-]\n
    tht: float, rotation of normal component about the first vector [rad]\n
    OUTPUTS:\n
    nx,ny,nz: float, components of the vector normal to first vector, with unit norm. [-]\n
    NOTE: of the two possible solutions, the one with ny<0 is returned
    """
    txOty = tx/ty
    nz= np.sin(theta)
    nx = np.sqrt( 1./ (1.+txOty**2))
    ny = -nx* txOty
    if ny>0: 
        nx *=-1
        ny *=-1
    return nx,ny,nz

def plot_legcable(cables_xyz,*args, D_cables=None,leg=False,ax=None):        
    """Plot leg and cables assembly.\n
    INPUT:\n
        cables_xyz: float(ncables,2,3), end A and end B coordinates in global CS.\n
        D_cables: float(ncables), cable diameters.\n
        leg: bool, whether or not the data is for the leg\n
        args: if provided, then final deflected configuration provided
        ax: axes, if provided it will be reused and final deflection assumed\n"""
    
    configs=1
    linetype0='solid'
    if len(args):
        cables_xyz2= args
        configs +=1
        linetype0='dotted'

    fig0= plt.figure()
    ax0 = plt.axes(projection='3d')
        
    n_cables= cables_xyz.shape[0]
    if len(D_cables)==0:
        D_cables=[np.nan]*n_cables
    for iconfig in range(configs):    
        mycables_xyz = cables_xyz
        linetype=linetype0
        if iconfig:
            mycables_xyz = args[0]
            linetype='solid'
        for icable,xyz in enumerate(mycables_xyz):
            Lg0=np.linalg.norm(xyz[1,:]-xyz[0,:])
            ax0.plot(xyz[:,0],xyz[:,1],xyz[:,2],label='D={:5.3f} m L= {:5.2f} m'.format(D_cables[icable],Lg0), linestyle=linetype)

        #lower_points = self.xyz[0,:]
                        
                        #0.5*self.D* [-sintht,costht], 
                      #- 0.5*self.D* [-sintht,costht]]

        #upper_point0 = self.Lg0 * [costht,sintht]  
        #upper_points = upper_point0*[ 1. + lower_points]

        #plt.fill(lower_points,upper_points, color=plt.cm.Dark2(loc_matidx[it]), label=cable)
        
        ax0.axis('equal')
        ax0.grid()
            #plt.title('Cable Geometry:  D={:5.3f} m L= {:5.2f} m '.format(*self.D,self.Lg0))
            #plt.plot(upper_points[0,:],upper_points[1,:],color='black',linestyle=':')
            #plt.legend(myhandles, mylabels)    
        ax0.set_xlabel('x_R[m]')
        ax0.set_ylabel('y_R[m]')
        ax0.set_zlabel('z_R[m]')
        ax0.legend()
        #plt.annotate('$\mu_s$({:3.2},{:3.2})= {:5.3e}'.format(*crit_points[ii,:], eps_z[var][ii]*10**6), xy=np.flip(crit_points[ii,:]), xycoords='data',\
        #                       xytext=(0, 10*(ii+1)*(-1)**i_surf), textcoords='offset points', arrowprops=dict(arrowstyle="->", connectionstyle="arc3"),  horizontalalignment='center', verticalalignment='bottom')
            
    plt.show()      
    return ax0

def print_results(prob,*args,post_opt=True, bline_yaml=None):
    """Print all results by not adding any args or specify a number of strings as in vars_preambles below to print those only.\n
    If bline_yaml is given as a full path to a file, a new yaml will be spat out as bline_yaml+"mod" """
    
    title = '               OPTIMIZATION RESULTS'
    if not(post_opt):
        title = '          PRE OPTIMIZATION RESULTS'
    
    print('_______________________________________________________')
    print(title)
    print('_______________________________________________________')

    print('D_LC= {:5.3f} m;  D_UC= {:5.3f} m  \n'.\
                format(prob.get_val('D_Cs.D_C')[0],  prob.get_val('D_Cs.D_C')[1]))

    print('sig0_LC= {:5.3f} MPa; sig0_UC= {:5.3f} MPa \n'.\
                format(prob.get_val('sig0_Cs.sig0_C')[0]/1.e6, prob.get_val('sig0_Cs.sig0_C')[1]/1.e6))
    
    print(('LC1.Nxyz= '+3*'{:5.3f} ' +'kN; UC1.Nxyz= '+3*'{:5.3f} ' +'kN \n').\
                format(*prob.get_val('LC1.Nxyz')/1.e3,*prob.get_val('UC1.Nxyz')/1.e3))

    print(('LC2.Nxyz= '+3*'{:5.3f} ' +'kN; UC2.Nxyz= '+3*'{:5.3f} ' +'kN\n').\
                format(*prob.get_val('LC2.Nxyz')/1.e3,*prob.get_val('UC2.Nxyz')/1.e3))

    print(('LC1.sig= {:5.3f} ' +'MPa; UC1.sig= {:5.3f} ' +'MPa \n').\
                format(*prob.get_val('LC1.sig')/1.e6,*prob.get_val('UC1.sig')/1.e6))

    print(('LC2.sig= {:5.3f} ' +'MPa; UC2.sig= {:5.3f} ' +'MPa \n').\
                format(*prob.get_val('LC2.sig')/1.e6,*prob.get_val('UC2.sig')/1.e6))

    print('Total cable mass (1 UC + 1 LC)= {:.3f} kg \n'.format(*prob.get_val('UC1LC1_mass.mass')))

    print(('Leg tip deflections x,y,z =  '+3*'{:5.3f} ' +'m \n').format(*prob.get_val('uvw_legtip')))
    print(('Leg tip twist             =  {:5.3f} deg \n').format(np.rad2deg(*prob.get_val('tht_legtip'))))
    
    print(('Leg root reactions x,y,z =  '+3*'{:5.3f} ' +'kN \n').format(*prob.get_val('LegCableBal.XYZ0')/1.e3))
    
    print('4ANSYS LC preload ={:10.5e} N\n'.format(prob.get_val('sig0_Cs.sig0_C')[0]*np.pi/4*prob.get_val('D_Cs.D_C')[0]**2))
    print('4ANSYS UC preload ={:10.5e} N\n'.format(prob.get_val('sig0_Cs.sig0_C')[1]*np.pi/4*prob.get_val('D_Cs.D_C')[1]**2))
    print('4ANSYS LC EA/L ={:10.5e} N/m\n'.format(LC_steel_def.E*np.pi/4*prob.get_val('D_Cs.D_C')[0]**2/np.linalg.norm(LC1_xyz0_def[1,:]-LC1_xyz0_def[0,:])))
    print('4ANSYS UC EA/L ={:10.5e} N/m\n'.format(UC_steel_def.E*np.pi/4*prob.get_val('D_Cs.D_C')[1]**2/np.linalg.norm(UC1_xyz0_def[1,:]-UC1_xyz0_def[0,:])))
 
    print('_______________________________________________________')
    print('_______________________________________________________')
    

    #Dump a yaml file after optimization
    if bline_yaml != None:
        thisdatetime = str(datetime.datetime.now())
        print('New yaml design processed by {:s} on {:s}\n'.format(os.path.basename(__file__), thisdatetime))
        new_yaml_file = Path(bline_yaml)

        des = prob.model
        if post_opt:
            des['comments'] += ' *** Post-Optimization *** '
            new_yaml_file = Path(new_yaml_file.parents[0],new_yaml_file.stem+"_optimized.yaml")
        else:
            new_yaml_file = Path(new_yaml_file.parents[0],new_yaml_file.stem+"_modified.yaml")    
        des['comments'] += '=> New yaml design processed by {:s} on {:s}\n'.format(os.path.basename(__file__), thisdatetime)
#        yaml = ryaml.YAML()
#        yaml.default_flow_style = None
#        yaml.width = float("inf")
#        yaml.indent(mapping=4, sequence=6, offset=3)
#        yaml.allow_unicode = False
       #TO DO figure out how to dump it nicely, because for now it does not work
        with open(new_yaml_file, "w", encoding="utf-8") as f2:  #, encoding="utf-8")
           yaml.dump(des, f2)

def read_inputs(xlsxfile,LC_xyzA_col="AG",UC_xyzA_col="Z",leg_col="F", leg_row=12, legmass_row=19, abc_col="J", abc_row=10, stem_col="B", stem_row=12):
    """This function opens up the geometry excel file and looks for cable xyz0, i.e., end A and end B coordinates.\n
    INPUT: LC_xyz0: float(2,3), lower cable endA, endB coordinates in absolute (global) CS. Units are [m]\n
           LC_xyzA_col: string(x), column name in the excel sheet for lower cable endA (=endB col) coordinates [-]\n
           LC_xyzA_row: int, row no. in the excel sheet for lower cable endA (=endB col) coordinates [-]\n
           UC_xyzA_col: string(x), column name in the excel sheet for upper cable endA (=endB col) coordinates [-]\n
           leg_col: string(x), column name in the excel sheet for leg D_L (assuming D_L,d_L,L_L stacked) [-]\n
           legmass_row: int, row no. in the excel sheet for leg wet mass  [-]\n
           stem_col: string(x), column name in the excel sheet for leg D_L (assuming D_S,d_S,L_S,draft_S stacked) [-]\n
           stem_row: int, row no. in the excel sheet for leg wet mass  [-]\n
        \n
    OUTPUT: LC_xyz0: float(2,3), lower cable endA, endB coordinates in absolute (global) CS. Units are [m]\n
            UC_xyz0: float(2,3), upper cable endA, endB coordinates in absolute (global) CS. Units are [m]\n
            [D_L,d_L,L_L0]: float(3), leg OD,ID and nominal length [m]\m \n
            M_Lwet: leg mass in water (weight as mass) [kg]\m \n
            abc: Geometry a,b,c paramaters, a=keel to leg-root hinge, c=leg-tip to leg-root vertical distance, b= top cable hinge to leg root vertical distance [m] \n

            """
    
    from openpyxl import load_workbook
  
    xl = load_workbook(xlsxfile,data_only=True)
    #Get the last sheet
    ws=xl[xl.sheetnames[-1]]  #Worksheet
    #lower cable coordinates
    LC_xyz0 = np.zeros([2,3]) #initialize
    UC_xyz0 = np.zeros([2,3]) #initialize
    LC_xyzA_col = "AG"
    UC_xyzA_col = "Z"
    LC_xyzA_row = UC_xyzA_row = 42
    LC_xyzB_row = UC_xyzB_row = 46
    LC_xyz0[0,:] = [float(cell[0].value) for cell in ws[LC_xyzA_col+str(LC_xyzA_row) +":" +LC_xyzA_col+str(LC_xyzA_row+2) ]]
    UC_xyz0[0,:] = [float(cell[0].value) for cell in ws[UC_xyzA_col+str(UC_xyzA_row) +":" +UC_xyzA_col+str(UC_xyzA_row+2) ]]
    LC_xyz0[1,:] = [float(cell[0].value) for cell in ws[LC_xyzA_col+str(LC_xyzB_row) +":" +LC_xyzA_col+str(LC_xyzB_row+2) ]] 
    UC_xyz0[1,:] = [float(cell[0].value) for cell in ws[UC_xyzA_col+str(UC_xyzB_row) +":" +UC_xyzA_col+str(UC_xyzB_row+2) ]] 
    [D_L,d_L,L_L0] = [float(cell[0].value) for cell in  ws[leg_col+str(leg_row) +":" + leg_col+str(leg_row+2)]]
    M_Lwet = ws[leg_col+str(legmass_row)].value *1.e3 #(it was in tonne)
    abc = [float(cell[0].value) for cell in  ws[abc_col+str(abc_row) +":" + abc_col+str(abc_row+2)] ]
    [D_S,d_S,L_S,draft_S]= [float(cell[0].value) for cell in ws[stem_col+str(stem_row) +":" + stem_col+str(stem_row+3)]] 
    
    return LC_xyz0, UC_xyz0, [D_L,d_L,L_L0],M_Lwet,abc,[D_S,d_S,L_S,draft_S]

def divide_cases(ncases, nprocs=2):
    """ Divide up load cases among available procs.\n

    INPUTS: ncases : int,   Number of load cases. \n
            nprocs : int,   Number of processors.\n
    OUTPUTS:  data: list of list of int,   Integer case numbers for each proc. \n
    """
    data = []
    for j in range(nprocs):
        data.append([])

    wrap = 0
    for j in range(ncases):
        idx = j - wrap
        if idx >= nprocs:
            idx = 0
            wrap = j

        data[idx].append(j)

    return data
#___________________________________________#


if __name__ == '__main__':
    """Here is an example on how to use this program"""

    #INPUTS  START
    baseline_yaml_def = Path("D:\RRD_ENGINEERING\PROJECTS\WE202402_ATLANTIS2\ANALYSIS\Optimization\LegCable.yaml")
    xls_input_def=Path("D:\\RRD_ENGINEERING\\PROJECTS\\WE202402_ATLANTIS2\\ANALYSIS\\USFLOWT_Geometry.xlsx")

    Fxyz_tip_def = np.array(([0.0, 100.e3, 6.e6, 0.,0.,0.],[0.0, 1.e6, 6.e6, 0.,0.,0.]))  # [n_loadcases X 6]
    D_L_def=3. # [m] leg OD
    L_L0_def=35. # [m] leg length
    L_Lxyz0_def = np.array([L_L0_def,0.,0.]) #components of leg vector root to tip in global CS
    M_L_def=267.e3 # [kg] leg mass
    zoff_def = [3.5,5.5] # [m] distance of lower and upper hinge from leg tip
    abc_def=[7.,29,0.] 

    #lower cable C1 end A and end B coordinates : see also USFLOWT_geometry.xlsx
    Rstem = 5; #[m] stem's OR 
    Rleg = 1.5; #[m] leg OR
    z_legroot_def = -17 #[m] leg root elevation
    zkeel= -24 #[m] keel elevation
    zifc=  +12. #[m] stem's ifc elevation  #Testing symmetric config
    cosa = np.cos(np.deg2rad(60.))
    sina = np.sin(np.deg2rad(60.))
    L_lh = L_L0_def - zoff_def[0] #length along leg-axis from stem to lower cable hinge
    LC1_xyz0_def = np.array([[Rstem*(cosa-1.), -Rstem*sina, zkeel  ],
                             [L_lh, -Rleg, L_Lxyz0_def[2]+z_legroot_def        ]])
    L_uh = L_L0_def - zoff_def[1] #length along leg-axis from stem to upper cable hinge
    UC1_xyz0_def = np.copy(LC1_xyz0_def)
    UC1_xyz0_def[0,2] = zifc
    UC1_xyz0_def[1,0] = L_uh
    
    D_c_def = [0.2,0.15] #lower,upper cable ODs [m]
    sig0_c_def = np.array([540.E6, 950.e6]) #lower, upper cable sigma0  constraints[Pa]
    sig0_bounds_def = np.array([50.E6, 1000.e6]) # sigma0  constraints[Pa]
    
    D_bounds_def = np.array([[0.05,0.5],[0.05,0.5]]) #lower and upper cable cOD min,max
    UC_steel_def=Material(matname='cable_steel', fp=1250.e6, E=195.e9, rho=7850.)
    LC_steel_def=Material(matname='cable_steel', fp=1250.e6, E=195.e9, rho=7850.)

    PSF_mat_def=1.2  #material yield PSF for cable

    opti = True
    #INPUTS END
    print('4ANSYS UC preload ={:} N'.format(sig0_c_def[1]*np.pi/4*D_c_def[1]**2))
    print('4ANSYS LC preload ={:} N'.format(sig0_c_def[0]*np.pi/4*D_c_def[0]**2))
    print('4ANSYS LC EA/L ={:} N/m'.format(LC_steel_def.E*np.pi/4*D_c_def[0]**2/np.linalg.norm(LC1_xyz0_def[1,:]-LC1_xyz0_def[0,:])))
    print('4ANSYS UC EA/L ={:} N/m'.format(UC_steel_def.E*np.pi/4*D_c_def[1]**2/np.linalg.norm(UC1_xyz0_def[1,:]-UC1_xyz0_def[0,:])))
    #Read inputs
    parser=argparse.ArgumentParser(description='SF_LegCable optimizer')
    
    parser.add_argument('--bline_yaml',    metavar='bline_yaml',             type=str,      help= 'Complete path to baseline yaml input file',    default=baseline_yaml_def)
    parser.add_argument('--xls_input',    metavar='xls_input',             type=str,      help= 'Complete path to xls input file',    default=xls_input_def)

    parser.add_argument('--D_c',       metavar='D_c',      type=float,    help= 'Upper, lower cable diameter fltarray(2)',      default=D_c_def)
    parser.add_argument('--sig0_c',    metavar='sig0_c',   type=float,    help= 'Lower, upper cable prestress guesses fltarray(2)',   default=sig0_c_def)
    parser.add_argument('--sig0_bounds', metavar='sig0_bounds',   type=float,    help= 'Lower, upper cable prestress limits fltarray(2)',   default=sig0_bounds_def)
    parser.add_argument('--abc',       metavar='abc',      type=float,    help= 'Geometry a,b,c paramaters, a=keel to leg-root hinge, c=leg-tip to leg-root vertical distance, b= top cable hinge to leg root vertical distance [m]',      default=abc_def)
    parser.add_argument('--PSF_mat',   metavar='PSF_mat',  type=float,    help= 'Yield PSF for the cables [-]',      default=PSF_mat_def)    
    parser.add_argument('--D_bounds',  metavar='D_bounds', type=float,    help= 'OD bounds for the cables [2,2] [m]',      default=D_bounds_def)
    parser.add_argument('--Fxyz_tip',  metavar='Fxyz_tip', type=float,    help= 'Forces and moments at Leg Tip',      default=Fxyz_tip_def)
    parser.add_argument('--D_L',       metavar='D_L',      type=float,    help= 'Leg OD [m]',      default=D_L_def)
    parser.add_argument('--L_L0',      metavar='L_L0',     type=float,    help= 'Leg length [m]',      default=L_L0_def)
    parser.add_argument('--L_Lxyz0',   metavar='L_Lxyz0',  type=float,    help= 'Leg vector components float(3) [m]',      default=L_Lxyz0_def)
    parser.add_argument('--z_legroot', metavar='z_legroot',type=float,    help= 'Leg root elevation [m]',      default=z_legroot_def)
    parser.add_argument('--M_L',       metavar='M_L',      type=float,    help= 'Leg mass [kg]',       default=M_L_def)
    parser.add_argument('--LC1_xyz0',  metavar='LC1_xyz0', type=float,    help= 'Lower C1 cable end A and end B coordinates [2,3] [m]',      default=LC1_xyz0_def)
    parser.add_argument('--UC1_xyz0',  metavar='UC1_xyz0', type=float,    help= 'Upper C1 cable end A and end B coordinates [2,3] [m]',      default=UC1_xyz0_def)

    parser.add_argument('--UCmat',     metavar='UCmat',    type=object,   help= 'Upper cable material [Material object]', default=UC_steel_def) 
    parser.add_argument('--LCmat',     metavar='LCmat',    type=object,   help= 'Lower cable material [Material object]', default=LC_steel_def) 
    parser.add_argument('--opti',      metavar='opti',     type=bool,     help= 'Whether or not to optimize',      default=opti)
    
    args=parser.parse_args()
    main(bline_yaml=args.bline_yaml, xls_input=args.xls_input, D_c=args.D_c, sig0_c=args.sig0_c, abc=args.abc, L_L0=args.L_L0,  L_Lxyz0=args.L_Lxyz0,z_legroot=args.z_legroot, D_L=args.D_L, M_L=args.M_L, Fxyz_tip=args.Fxyz_tip, \
         LC1_xyz0=args.LC1_xyz0, UC1_xyz0=args.UC1_xyz0, PSF_mat=args.PSF_mat, D_bounds=args.D_bounds, sig0_bounds=args.sig0_bounds, LCmat=args.LCmat, UCmat=args.UCmat,  opti=args.opti)

